#!/usr/bin/env python3
"""
Iowa Election Data Processor
============================
This script processes Iowa election Excel files to:
  1. Keep only major races (President, Governor, US Senate, US House, State Senate, State House)
  2. Keep only the "Total" vote columns (remove Absentee and Polling breakdowns)
  3. Clean up column names (remove " Total" suffix)

Designed to be readable for students with ~1.5 semesters of programming experience.

Required library: openpyxl (install with: pip install openpyxl)
"""

import openpyxl
from openpyxl import Workbook
import glob
import os

# ============================================================================
# STEP 1: DEFINE WHICH RACES TO KEEP
# ============================================================================

def is_race_we_want(race_name):
    """
    Check if a race should be kept in our filtered data.

    We want to keep major political races and remove judicial retention votes.

    Parameters:
        race_name: The name of the race (e.g., "President and Vice President")

    Returns:
        True if we should keep this race, False if we should remove it
    """
    # Handle empty or invalid race names
    if race_name is None:
        return False
    if not isinstance(race_name, str):
        return False

    # Convert to lowercase for easier matching
    race_lower = race_name.lower()

    # List of keywords that identify races we want to KEEP
    races_to_keep = [
        'president',
        'governor',
        'u.s. senator',
        'us senator',
        'u.s. rep',
        'us rep',
        'state senator',
        'state rep',
    ]

    # Check if any of our keywords appear in the race name
    for keyword in races_to_keep:
        if keyword in race_lower:
            return True

    # If no keywords matched, we don't want this race
    return False


# ============================================================================
# STEP 2: IDENTIFY WHICH COLUMNS TO KEEP
# ============================================================================

def get_columns_to_keep(header_row):
    """
    Determine which columns to keep from the spreadsheet.

    We keep:
      - RaceTitle, CandidateName, PoliticalPartyName (metadata columns)
      - Any column ending in " Total" (renamed to remove that suffix)

    We remove:
      - Absentee columns
      - Polling columns
      - County total columns (no hyphen in name, like just "Polk")

    Parameters:
        header_row: List of column headers from row 1 of the spreadsheet

    Returns:
        List of tuples: (original_column_index, new_column_name)
    """
    columns_to_keep = []

    # Metadata columns we always keep (using their original names)
    metadata_columns = ['RaceTitle', 'CandidateName', 'PoliticalPartyName']

    # Go through each column header
    for column_index, header in enumerate(header_row):

        # Skip empty headers
        if header is None:
            continue

        # Clean up whitespace
        header_clean = header.strip()

        # Check if it's a metadata column
        if header_clean in metadata_columns:
            columns_to_keep.append((column_index, header_clean))

        # Check if it's a "Total" column (these are the vote totals we want)
        elif header_clean.endswith(' Total'):
            # Remove the " Total" suffix to get just the precinct name
            # " Total" is 6 characters, so we slice off the last 6
            precinct_name = header_clean[:-6]

            # Only keep precinct totals (they have a hyphen like "Polk-Precinct1")
            # Skip county totals (they have no hyphen, like just "Polk")
            if '-' in precinct_name:
                columns_to_keep.append((column_index, precinct_name))

    return columns_to_keep


# ============================================================================
# STEP 3: MAIN PROCESSING FUNCTION
# ============================================================================

def process_election_file(input_filename, output_filename):
    """
    Process an Iowa election Excel file.

    Parameters:
        input_filename: Path to the original Excel file
        output_filename: Path where the cleaned file should be saved
    """
    print(f"Processing: {input_filename}")
    print("-" * 50)

    # --- Load the source file ---
    print("Loading file...")
    source_workbook = openpyxl.load_workbook(input_filename, data_only=True)
    source_sheet = source_workbook.active

    original_rows = source_sheet.max_row
    original_cols = source_sheet.max_column
    print(f"Original size: {original_rows} rows x {original_cols} columns")

    # --- Get the header row ---
    header_row = [cell.value for cell in source_sheet[1]]

    # --- Determine which columns to keep ---
    columns_to_keep = get_columns_to_keep(header_row)
    print(f"Columns to keep: {len(columns_to_keep)}")

    # --- Create the new workbook ---
    new_workbook = Workbook()
    new_sheet = new_workbook.active
    new_sheet.title = "Election Data"

    # --- Write the new header row ---
    new_headers = [name for (index, name) in columns_to_keep]
    new_sheet.append(new_headers)

    # --- Process each data row ---
    rows_kept = 0
    rows_removed = 0

    # Start at row 2 (row 1 is headers)
    for row in source_sheet.iter_rows(min_row=2, values_only=True):

        # Get the race name (first column)
        race_name = row[0]

        # Check if we want to keep this race
        if is_race_we_want(race_name):
            # Extract only the columns we want to keep
            new_row = [row[index] for (index, name) in columns_to_keep]
            new_sheet.append(new_row)
            rows_kept += 1
        else:
            rows_removed += 1

    # --- Save the new file ---
    print(f"Rows kept: {rows_kept}")
    print(f"Rows removed: {rows_removed}")

    new_workbook.save(output_filename)
    source_workbook.close()

    print(f"Saved to: {output_filename}")
    print(f"Final size: {rows_kept + 1} rows x {len(columns_to_keep)} columns")
    print()


# ============================================================================
# STEP 4: MERGE COUNTY FILES (for 2016 data)
# ============================================================================

def merge_county_files(county_files_pattern, output_filename):
    """
    Merge multiple county election files into one statewide file.

    Each county file has the same rows (races/candidates) but different
    precinct columns. We combine them horizontally.

    Parameters:
        county_files_pattern: Glob pattern like "*_2016_totals.xlsx"
        output_filename: Where to save the merged file
    """
    print("=" * 60)
    print("MERGING COUNTY FILES")
    print("=" * 60)

    # --- Find all county files ---
    county_files = sorted(glob.glob(county_files_pattern))
    print(f"Found {len(county_files)} county files to merge")

    if len(county_files) == 0:
        print("ERROR: No files found matching pattern!")
        return

    # --- Load the first file to get the base structure ---
    print(f"\nLoading first file as template: {os.path.basename(county_files[0])}")
    first_wb = openpyxl.load_workbook(county_files[0], data_only=True)
    first_ws = first_wb.active

    # Get the "key" columns (RaceTitle, CandidateName) that identify each row
    # These should be the same in every county file
    base_rows = []
    for row in first_ws.iter_rows(min_row=2, max_col=2, values_only=True):
        base_rows.append((row[0], row[1]))  # (RaceTitle, CandidateName)

    print(f"Base structure: {len(base_rows)} candidate rows")

    # Get the first file's headers and precinct columns
    first_headers = [cell.value for cell in first_ws[1]]

    # Separate metadata columns from precinct columns
    metadata_cols = ['RaceTitle', 'CandidateName', 'PoliticalPartyName']

    # Start building the merged data
    # merged_headers will have: RaceTitle, CandidateName, then all precincts
    merged_headers = ['RaceTitle', 'CandidateName']

    # merged_data[row_index] = [RaceTitle, CandidateName, precinct1_votes, precinct2_votes, ...]
    merged_data = []
    for row in first_ws.iter_rows(min_row=2, values_only=True):
        merged_data.append([row[0], row[1]])  # Start with RaceTitle, CandidateName

    first_wb.close()

    # --- Process each county file ---
    total_precincts = 0
    problems_found = []

    for file_path in county_files:
        county_name = os.path.basename(file_path).replace("_2016_totals.xlsx", "")
        print(f"  Adding: {county_name}...", end=" ")

        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active

        # Get headers
        headers = [cell.value for cell in ws[1]]

        # Find precinct columns (not metadata)
        precinct_indices = []
        for i, h in enumerate(headers):
            if h and h.strip() not in metadata_cols:
                precinct_indices.append(i)
                merged_headers.append(h.strip())

        print(f"{len(precinct_indices)} precincts")
        total_precincts += len(precinct_indices)

        # Get data rows and verify they match the base structure
        row_index = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            race_title = row[0]
            candidate = row[1]

            # Check if this row matches what we expect
            if row_index < len(base_rows):
                expected_race, expected_candidate = base_rows[row_index]
                if race_title != expected_race or candidate != expected_candidate:
                    problems_found.append(
                        f"Row mismatch in {county_name}: expected ({expected_race}, {expected_candidate}), "
                        f"got ({race_title}, {candidate})"
                    )

            # Add this county's precinct votes to the merged data
            precinct_votes = [row[i] for i in precinct_indices]
            merged_data[row_index].extend(precinct_votes)

            row_index += 1

        wb.close()

    print(f"\nTotal precincts across all counties: {total_precincts}")

    # --- Verify the merged data ---
    print("\nRunning verification checks...")

    # Check 1: All rows should have the same number of columns
    expected_cols = 2 + total_precincts  # RaceTitle + CandidateName + all precincts
    for i, row in enumerate(merged_data):
        if len(row) != expected_cols:
            problems_found.append(
                f"Row {i+2} has {len(row)} columns, expected {expected_cols}"
            )

    # Check 2: Look for votes in wrong precincts
    # For district races, most precincts should have 0 or None
    # Statewide races should have votes in most precincts
    print("  Checking vote patterns...")

    statewide_races = ['president', 'governor', 'u.s. senator', 'us senator']

    for i, row in enumerate(merged_data):
        race_title = row[0].lower() if row[0] else ""
        candidate = row[1]
        precinct_votes = row[2:]  # Everything after RaceTitle, CandidateName

        # Count precincts with non-zero votes
        precincts_with_votes = 0
        for v in precinct_votes:
            try:
                if v and int(v) > 0:
                    precincts_with_votes += 1
            except (ValueError, TypeError):
                pass

        # For statewide races, we expect votes in most precincts
        is_statewide = any(pattern in race_title for pattern in statewide_races)

        if is_statewide and precincts_with_votes < total_precincts * 0.5:
            # This might be a minor candidate with few votes, which is OK
            pass

        # For district races, having votes in ALL precincts would be suspicious
        if not is_statewide and precincts_with_votes == total_precincts:
            problems_found.append(
                f"Suspicious: '{race_title}' / '{candidate}' has votes in ALL precincts"
            )

    # --- Report problems ---
    if problems_found:
        print(f"\n*** FOUND {len(problems_found)} POTENTIAL PROBLEMS ***")
        for p in problems_found[:10]:  # Show first 10
            print(f"  - {p}")
        if len(problems_found) > 10:
            print(f"  ... and {len(problems_found) - 10} more")
    else:
        print("  No problems found!")

    # --- Save the merged file ---
    print(f"\nSaving merged file...")

    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.title = "Statewide 2016"

    # Write headers
    new_ws.append(merged_headers)

    # Write data rows
    for row in merged_data:
        new_ws.append(row)

    new_wb.save(output_filename)

    print(f"Saved: {output_filename}")
    print(f"Final size: {len(merged_data) + 1} rows x {len(merged_headers)} columns")
    print()


# ============================================================================
# STEP 5: RUN THE SCRIPT
# ============================================================================

if __name__ == "__main__":
    """
    This section runs when you execute the script directly.
    Modify the filenames below to process different files.
    """

    # ----- PROCESS 2014 STATEWIDE FILE -----
    # (Already in statewide format, just needs filtering)
    print("\n" + "=" * 60)
    print("PROCESSING 2014 DATA")
    print("=" * 60)

    process_election_file(
        input_filename="statewide_2014.xlsx",
        output_filename="statewide_2014_cleaned.xlsx"
    )

    # ----- MERGE 2016 COUNTY FILES -----
    # (99 county files need to be merged into one statewide file)
    # NOTE: Run this AFTER the county files have been processed
    #       to create the *_2016_totals.xlsx files

    print("\n" + "=" * 60)
    print("PROCESSING 2016 DATA")
    print("=" * 60)

    merge_county_files(
        county_files_pattern="*_2016_totals.xlsx",
        output_filename="statewide_2016_merged.xlsx"
    )

    # ----- FUTURE: 2018 and 2020 -----
    # Add processing for these years when data is available
    # process_election_file("statewide_2018.xlsx", "statewide_2018_cleaned.xlsx")
    # process_election_file("statewide_2020.xlsx", "statewide_2020_cleaned.xlsx")
